#!/usr/bin/env python
from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
import tkinter as tk
from tkinter import ttk

from dispatch.aggregate_warnings import aggregate_warnings, gather_valid_names


class AssignmentApp(tk.Tk):
    """Simple drag-and-drop GUI to map unknown names to known technicians."""

    def __init__(self, unknown: Counter[str], valid: list[str], liste: Path) -> None:
        super().__init__()
        self.title("Techniker-Zuordnung")

        self._drag_item: str | None = None
        self.mappings: dict[str, str] = {}
        self.liste_path = liste

        # Unknown names list
        left = ttk.Frame(self)
        left.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        ttk.Label(left, text="Unbekannte Namen").pack()
        self.unknown = tk.Listbox(left)
        self.unknown.pack(fill="both", expand=True)
        for name, count in unknown.items():
            self.unknown.insert("end", f"{name} ({count})")
        self.unknown.bind("<ButtonPress-1>", self._start_drag)

        # Known names list
        middle = ttk.Frame(self)
        middle.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        ttk.Label(middle, text="Bekannte Techniker").pack()
        self.valid = tk.Listbox(middle)
        self.valid.pack(fill="both", expand=True)
        for name in sorted(set(valid)):
            self.valid.insert("end", name)
        self.valid.bind("<ButtonRelease-1>", self._on_drop)

        # Mappings display
        right = ttk.Frame(self)
        right.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        ttk.Label(right, text="Zuordnungen").pack()
        self.result = tk.Listbox(right)
        self.result.pack(fill="both", expand=True)

        ttk.Button(self, text="Export", command=self._export).pack(side="bottom", pady=5)

    def _start_drag(self, event: tk.Event) -> None:
        index = self.unknown.nearest(event.y)
        if index >= 0:
            self._drag_item = self.unknown.get(index)

    def _on_drop(self, event: tk.Event) -> None:
        if not self._drag_item:
            return
        unknown_name = self._drag_item.split(" (", 1)[0]
        index = self.valid.nearest(event.y)
        if index >= 0:
            known_name = self.valid.get(index)
            self.mappings[unknown_name] = known_name
            self.result.insert("end", f"{unknown_name} -> {known_name}")
            # remove from unknown list
            items = self.unknown.get(0, "end")
            self.unknown.delete(0, "end")
            for item in items:
                if not item.startswith(unknown_name + " "):
                    self.unknown.insert("end", item)
        self._drag_item = None

    def _export(self) -> None:
        """Export the technician mappings to a CSV file and Liste.xlsx."""
        export_path = Path(__file__).resolve().parent / "techniker_export.csv"
        with export_path.open("w", encoding="utf-8") as fh:
            for unknown, known in self.mappings.items():
                fh.write(f"{unknown},{known}\n")
        print(f"Exportiert nach {export_path}")

        try:
            from openpyxl import load_workbook  # type: ignore
            from openpyxl.utils.exceptions import InvalidFileException  # type: ignore

            wb = load_workbook(self.liste_path)
            sheet_name = "Zuordnungen"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Unbekannt", "Bekannt"])
            for unknown, known in self.mappings.items():
                ws.append([unknown, known])
            wb.save(self.liste_path)
            wb.close()
            print(f"Liste aktualisiert: {self.liste_path}")
        except ImportError:
            print("Konnte Liste.xlsx nicht aktualisieren: openpyxl ist nicht installiert")
        except InvalidFileException as exc:
            print(f"Konnte Liste.xlsx nicht aktualisieren: {exc}")
        except Exception as exc:
            print(f"Konnte Liste.xlsx nicht aktualisieren: {exc}")
            raise

        self.destroy()


def main(argv: list[str] | None = None) -> None:
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Interaktive Zuordnung unbekannter Techniker"
    )
    parser.add_argument(
        "report_dir",
        nargs="?",
        type=Path,
        default=base_dir / "data",
        help="Verzeichnis mit Tagesberichten",
    )
    parser.add_argument(
        "--liste",
        type=Path,
        default=base_dir / "data" / "Liste.xlsx",
        help="Pfad zur Liste.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default="Technikernamen",
        help="Name des Tabellenblatts mit Technikern",
    )
    args = parser.parse_args(argv)

    try:
        valid = gather_valid_names(args.liste, sheet_name=args.sheet)
        unknown = aggregate_warnings(args.report_dir, valid)
    except RuntimeError as exc:  # missing dependency like openpyxl
        print(exc)
        print("Install required dependencies with: pip install openpyxl")
        return

    app = AssignmentApp(unknown, valid, args.liste)
    app.mainloop()


if __name__ == "__main__":  # pragma: no cover - GUI entry point
    main()
