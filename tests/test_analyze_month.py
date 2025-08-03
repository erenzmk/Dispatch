import csv
import datetime as dt
from pathlib import Path
import sys

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from analyze_month import analyze_month


def _create_report(path: Path, names: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = "Employee ID"
    ws["B5"] = "Employee Name"
    ws["C5"] = "Open Date Time"
    row = 6
    for name in names:
        ws.cell(row=row, column=1).value = 1
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=3).value = dt.datetime(2025, 6, 30)
        row += 1
    wb.save(path)


def test_analyze_month_detects_missing_and_extra(tmp_path):
    month_dir = tmp_path / "Foo_25"
    month_dir.mkdir()

    day1 = month_dir / "01.07"
    day1.mkdir()
    _create_report(day1 / "m7.xlsx", ["Alice"])
    _create_report(day1 / "e19.xlsx", ["Alice"])

    day2 = month_dir / "02.07"
    day2.mkdir()
    _create_report(day2 / "m7.xlsx", ["Charlie"])
    _create_report(day2 / "e19.xlsx", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Foo_25"
    ws["A1"] = "name"
    ws["A2"] = "Alice"
    ws["A3"] = "Bob"
    liste_path = tmp_path / "Liste.xlsx"
    wb.save(liste_path)

    out_csv = tmp_path / "out.csv"
    analyze_month(month_dir, liste_path, out_csv)

    with out_csv.open() as fh:
        rows = list(csv.reader(fh))

    assert rows[0] == ["category", "technician"]
    assert ["no_calls", "Bob"] in rows[1:]
    assert ["region_mismatch", "Charlie"] in rows[1:]
