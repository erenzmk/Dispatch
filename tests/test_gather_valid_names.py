import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dispatch.aggregate_warnings import gather_valid_names


def test_gather_valid_names_reads_sheet_and_deduplicates(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Technikernamen"
    ws.append(["Technikername", "PUOOS"])
    ws.append(["Alice", "Ali"])
    ws.append(["Bob", None])
    ws.append(["Alice", ""])
    wb.save(tmp_path / "Liste.xlsx")
    wb.close()

    names = gather_valid_names(tmp_path / "Liste.xlsx")
    assert names == ["Ali", "Alice", "Bob"]


def test_gather_valid_names_autodetects_sheet(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Technik Januar"
    ws.append(["Technikername"])
    ws.append(["Foo"])
    ws2 = wb.create_sheet("Technikernamen + PUDO")
    ws2.append(["Technikername"])
    ws2.append(["Eva"])
    wb.save(tmp_path / "Liste.xlsx")
    wb.close()

    assert gather_valid_names(tmp_path / "Liste.xlsx") == ["Eva"]


def test_gather_valid_names_raises_when_missing(tmp_path):
    wb = Workbook()
    wb.save(tmp_path / "Liste.xlsx")
    wb.close()

    with pytest.raises(ValueError):
        gather_valid_names(tmp_path / "Liste.xlsx")
