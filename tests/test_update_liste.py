import datetime as dt
from pathlib import Path

from openpyxl import Workbook, load_workbook

from process_reports import update_liste
from name_aliases import refresh_alias_map


def test_update_liste(tmp_path: Path):
    refresh_alias_map()
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=3, column=1, value="Bob")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 3, "new": 1, "old": 2}}
    evening = {"Alice": {"total": 1, "new": 0, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning, evening)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]

    assert ws2.cell(row=2, column=8).value == 2
    assert ws2.cell(row=2, column=9).value == 3
    assert ws2.cell(row=2, column=10).value == 2
    assert ws2.cell(row=2, column=11).value == 1
    assert ws2.cell(row=3, column=8).value is None
    assert ws2.cell(row=3, column=9).value is None
    assert ws2.cell(row=3, column=10).value is None
    assert ws2.cell(row=3, column=11).value is None


def test_update_liste_missing_evening(tmp_path: Path):
    refresh_alias_map()
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=2, column=1, value="Alice")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {
        "Alice": {"total": 4, "new": 1, "old": 3},
        "Bob": {"total": 2, "new": 0, "old": 2},
    }
    evening = {}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning, evening)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]

    assert ws2.cell(row=2, column=8).value == 4
    assert ws2.cell(row=2, column=9).value == 4
    assert ws2.cell(row=2, column=10).value == 3
    assert ws2.cell(row=2, column=11).value == 1
    assert ws2.cell(row=3, column=1).value == "Bob"
    assert ws2.cell(row=3, column=8).value == 2
    assert ws2.cell(row=3, column=9).value == 2
    assert ws2.cell(row=3, column=10).value == 2
    assert ws2.cell(row=3, column=11).value == 0
    assert ws2.max_row == 3


def test_update_liste_evening_extra_technicians(tmp_path: Path):
    refresh_alias_map()
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=2, column=1, value="Alice")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {
        "Alice": {"total": 5, "new": 2, "old": 3},
        "Bob": {"total": 3, "new": 1, "old": 2},
    }
    evening = {
        "Alice": {"total": 1, "new": 0, "old": 1},
        "Charlie": {"total": 2, "new": 1, "old": 1},
    }

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning, evening)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]

    assert ws2.cell(row=2, column=8).value == 4
    assert ws2.cell(row=2, column=9).value == 5
    assert ws2.cell(row=2, column=10).value == 3
    assert ws2.cell(row=2, column=11).value == 2
    assert ws2.cell(row=3, column=1).value == "Bob"
    assert ws2.cell(row=3, column=8).value == 3
    assert ws2.cell(row=3, column=9).value == 3
    assert ws2.cell(row=3, column=10).value == 2
    assert ws2.cell(row=3, column=11).value == 1
    assert ws2.max_row == 3
    names = [ws2.cell(row=r, column=1).value for r in range(2, ws2.max_row + 1)]
    assert "Charlie" not in names
