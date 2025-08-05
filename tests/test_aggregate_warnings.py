import datetime as dt
import logging
from pathlib import Path
import sys

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dispatch.aggregate_warnings import aggregate_warnings

def _write_report(path: Path, tech: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = "Employee ID"
    ws["B5"] = "Employee Name"
    ws["C5"] = "Open Date Time"
    ws["A6"] = 1
    ws["B6"] = tech
    ws["C6"] = dt.datetime(2025, 6, 30)
    wb.save(path)

def test_aggregate_warnings_does_not_touch_global_logging(tmp_path):
    report_dir = tmp_path / "reports"
    report_dir.mkdir()
    _write_report(report_dir / "report.xlsx", "Bob")

    root_logger = logging.getLogger()
    process_logger = logging.getLogger("dispatch.process_reports")
    root_level = root_logger.level
    process_level = process_logger.level
    root_handlers = list(root_logger.handlers)
    process_handlers = list(process_logger.handlers)

    counter = aggregate_warnings(report_dir, ["Alice"])

    assert counter == {"Bob": 1}
    assert root_logger.level == root_level
    assert list(root_logger.handlers) == root_handlers
    assert process_logger.level == process_level
    assert list(process_logger.handlers) == process_handlers


def test_aggregate_warnings_counts_multiple_reports(tmp_path):
    report_dir = tmp_path / "reports"
    report_dir.mkdir()
    _write_report(report_dir / "r1.xlsx", "Bob")
    _write_report(report_dir / "r2.xlsx", "Charlie")
    _write_report(report_dir / "r3.xlsx", "Bob")

    counter = aggregate_warnings(report_dir, ["Alice"])

    assert counter == {"Bob": 2, "Charlie": 1}
