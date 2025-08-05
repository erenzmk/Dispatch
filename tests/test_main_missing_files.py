import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
import datetime as dt

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dispatch.process_reports import main


def create_liste(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    wb.save(path)


def test_main_missing_morning_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    liste = tmp_path / "Liste.xlsx"
    create_liste(liste)

    day_dir = tmp_path / "01.07"
    day_dir.mkdir()

    # create an evening file to ensure only morning is missing
    wb = Workbook()
    wb.save(day_dir / "evening19.xlsx")

    monkeypatch.setattr(sys, "argv", ["process_reports.py", str(day_dir), str(liste)])
    with pytest.raises(FileNotFoundError) as excinfo:
        main()
    assert "Morning report" in str(excinfo.value)
    assert str(day_dir) in str(excinfo.value)


def test_main_no_evening_file_ok(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    liste = tmp_path / "Liste.xlsx"
    create_liste(liste)

    day_dir = tmp_path / "01.07"
    day_dir.mkdir()

    # create a morning file but no evening file
    wb = Workbook()
    wb.save(day_dir / "morning7.xlsx")

    monkeypatch.setattr(sys, "argv", ["process_reports.py", str(day_dir), str(liste)])

    def fake_load_calls(path, valid_names=None):
        return dt.date(2025, 7, 1), {}, []

    def fake_update_liste(liste_path, month_sheet, target_date, morning_summary):
        pass

    monkeypatch.setattr("dispatch.process_reports.load_calls", fake_load_calls)
    monkeypatch.setattr("dispatch.process_reports.update_liste", fake_update_liste)

    # Should not raise even if evening file is missing
    main()
