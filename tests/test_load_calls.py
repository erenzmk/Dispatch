import datetime as dt
from pathlib import Path
import sys
import os

import pytest
from openpyxl import Workbook
import logging

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from process_reports import load_calls


def test_load_calls_selects_correct_sheet(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1["A1"] = "not the sheet"

    ws2 = wb.create_sheet("Report")
    ws2["A2"] = dt.datetime(2025, 7, 1)
    ws2["A5"] = "Employee ID"
    ws2["B5"] = "Employee Name"
    ws2["C5"] = "Open Date Time"
    ws2["A6"] = 1
    ws2["B6"] = "Alice"
    ws2["C6"] = dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Alice": {"total": 1, "new": 1, "old": 0}}


def test_load_calls_handles_header_variations(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = " employee id "
    ws["B5"] = "EMPLOYEE NAME "
    ws["C5"] = " open date time "
    ws["A6"] = 1
    ws["B6"] = "Alice"
    ws["C6"] = dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Alice": {"total": 1, "new": 1, "old": 0}}


def test_load_calls_missing_required_columns(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = "Employee ID"
    ws["B5"] = "Employee Name"  # missing Open Date Time
    ws["A6"] = 1
    ws["B6"] = "Alice"

    path = tmp_path / "report.xlsx"
    wb.save(path)

    with pytest.raises(ValueError) as exc:
        load_calls(path)
    assert "Open Date Time" in str(exc.value)


def test_load_calls_does_not_leak_file_handles(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = "Employee ID"
    ws["B5"] = "Employee Name"
    ws["C5"] = "Open Date Time"
    ws["A6"] = 1
    ws["B6"] = "Alice"
    ws["C6"] = dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    def fd_count() -> int:
        return len(os.listdir("/proc/self/fd"))

    baseline = fd_count()
    for _ in range(3):
        load_calls(path)
        assert fd_count() == baseline


def test_load_calls_logs_warning_for_unknown_name(tmp_path, caplog):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = "Employee ID"
    ws["B5"] = "Employee Name"
    ws["C5"] = "Open Date Time"
    ws["A6"] = 1
    ws["B6"] = "Someone"
    ws["C6"] = dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    valid_names = ["Alice"]
    with caplog.at_level(logging.WARNING):
        load_calls(path, valid_names)
    assert any("Someone" in rec.message for rec in caplog.records)
