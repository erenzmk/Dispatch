import datetime as dt
import sys
from pathlib import Path
import logging

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dispatch.process_reports import process_month


def create_liste(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    wb.save(path)


def test_process_month_multiple_days(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
) -> None:
    month_dir = tmp_path / "Juli_25"
    day1 = month_dir / "01.07"
    day2 = month_dir / "02.07"

    for d in (day1, day2):
        d.mkdir(parents=True)
        wb = Workbook()
        wb.save(d / "m7.xlsx")

    liste = tmp_path / "Liste.xlsx"
    create_liste(liste)

    def fake_load_calls(path, valid_names=None):
        day = int(Path(path).parent.name.split(".")[0])
        return dt.date(2025, 7, day), {}

    calls: list[dt.date] = []

    def fake_update_liste(liste_path, month_sheet, target_date, morning_summary):
        calls.append(target_date)

    monkeypatch.setattr("dispatch.process_reports.load_calls", fake_load_calls)
    monkeypatch.setattr("dispatch.process_reports.update_liste", fake_update_liste)

    with caplog.at_level(logging.WARNING):
        process_month(month_dir, liste, log_file=None)

    assert calls == [dt.date(2025, 7, 1), dt.date(2025, 7, 2)]


def test_process_month_logging(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    month_dir = tmp_path / "Juli_25"
    day1 = month_dir / "01.07"
    day1.mkdir(parents=True)
    wb = Workbook()
    wb.save(day1 / "m7.xlsx")

    liste = tmp_path / "Liste.xlsx"
    create_liste(liste)

    def fake_load_calls(path, valid_names=None):
        return dt.date(2025, 7, 1), {}

    def fake_update_liste(liste_path, month_sheet, target_date, morning_summary):
        pass

    monkeypatch.setattr("dispatch.process_reports.load_calls", fake_load_calls)
    monkeypatch.setattr("dispatch.process_reports.update_liste", fake_update_liste)

    log_file = tmp_path / "run.log"
    process_month(month_dir, liste, log_file)

    content = log_file.read_text(encoding="utf-8")
    assert "Verarbeite 01.07" in content
    assert "Monatsverarbeitung abgeschlossen" in content
