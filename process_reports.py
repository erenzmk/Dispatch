"""
Automate dispatch call summaries.

This script processes daily technician call reports. For each day directory
(e.g. ``Juli_25/01.07``) it expects a morning file named with ``7`` and an
evening file named with ``19``. The morning report is used to determine for
each technician the number of calls in total as well as how many are "new"
(``Open Date Time`` equals the previous business day) or "old".  The evening
report lists the still open calls which allows the script to derive how many
were completed during the day.

The aggregated values are written into ``Liste.xlsx`` on the sheet for the
corresponding month. The workbook contains weekly column blocks consisting of
13 columns:

    name, date, weekday, pudo, pickup time, valid, info, pre-closed,
    total calls, old calls, new calls, details, mails

Blocks are repeated for each week and separated by an empty column.  The first
block starts at column ``A``.  To update values for a given date ``d`` the
block index is ``(d.day-1)//7`` and the starting column is ``1 + index*14``.

The script requires :mod:`openpyxl` for reading and writing Excel files.
"""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Dict

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - missing dependency
    raise SystemExit("openpyxl is required: {}".format(exc))


HEADER_MARKER = "Employee ID"
PREV_DAY_MAP = {
    0: "Montag",
    1: "Dienstag",
    2: "Mittwoch",
    3: "Donnerstag",
    4: "Freitag",
    5: "Samstag",
    6: "Sonntag",
}

# Mapping of month numbers to German month names used in ``Liste.xlsx``
MONTH_MAP = {
    1: "Januar",
    2: "Februar",
    3: "März",
    4: "April",
    5: "Mai",
    6: "Juni",
    7: "Juli",
    8: "August",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Dezember",
}


def excel_to_date(value):
    """Convert an Excel serial or datetime to a :class:`datetime.date`."""
    if isinstance(value, dt.datetime):
        return value.date()
    origin = dt.date(1899, 12, 30)
    return (origin + dt.timedelta(days=float(value)))


def prev_business_day(day: dt.date) -> dt.date:
    """Return the previous business day (skip weekends)."""
    day -= dt.timedelta(days=1)
    while day.weekday() >= 5:  # 5=Sat,6=Sun
        day -= dt.timedelta(days=1)
    return day


def load_calls(path: Path) -> Dict[str, Dict[str, int]]:
    """Load a call report and summarise per technician."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    header_row = None
    header_row_idx = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True),
@@ -94,69 +110,71 @@ def load_calls(path: Path) -> Dict[str, Dict[str, int]]:
    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=20, values_only=True), 1
    ):
        if row and row[0] == HEADER_MARKER:
            header_row = list(row)
            header_row_idx = idx
            break
    if header_row is None:
        raise ValueError("Header row not found in report")

    name_idx = header_row.index("Employee Name")
    open_idx = header_row.index("Open Date Time")

    target_cell = ws.cell(row=2, column=1).value
    target_date = excel_to_date(target_cell)
    prev_day = prev_business_day(target_date)

    summary: Dict[str, Dict[str, int]] = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row and isinstance(row[0], str) and row[0] == HEADER_MARKER:
            continue
        if not row or row[name_idx] in (None, ""):
            continue
        tech = str(row[name_idx]).strip()
        open_date = excel_to_date(row[open_idx])
        data = summary.setdefault(tech, {"total": 0, "new": 0, "old": 0})
        data["total"] += 1
        if open_date == prev_day:
            data["new"] += 1
        else:
            data["old"] += 1
    return target_date, summary


def update_liste(
    liste: Path,
    month_sheet: str,
    day: dt.date,
    morning: Dict[str, Dict[str, int]],
    evening: Dict[str, Dict[str, int]],
):
    """Write aggregated values into the ``Liste.xlsx`` workbook."""
    wb = load_workbook(liste)
    if month_sheet not in wb.sheetnames:
        raise KeyError(f"Worksheet {month_sheet} does not exist in {liste}")
    ws = wb[month_sheet]

    week_index = (day.day - 1) // 7
    start_col = 1 + week_index * 14

    for row in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=1)
        tech = str(name_cell.value).strip() if name_cell.value else None
        if not tech or tech not in morning:
            continue
        day_data = morning[tech]
        eve_total = evening.get(tech, {}).get("total", 0)
        closed = day_data["total"] - eve_total
        ws.cell(row=row, column=start_col + 1).value = day.toordinal() + 693594
        ws.cell(row=row, column=start_col + 2).value = PREV_DAY_MAP[day.weekday()]
        ws.cell(row=row, column=start_col + 7).value = closed
        ws.cell(row=row, column=start_col + 8).value = day_data["total"]
        ws.cell(row=row, column=start_col + 9).value = day_data["old"]
        ws.cell(row=row, column=start_col + 10).value = day_data["new"]
    wb.save(liste)


def main():
    parser = argparse.ArgumentParser(description="Process dispatch reports")
    parser.add_argument("day_dir", type=Path,
                        help="Directory containing daily reports (e.g. Juli_25/01.07)")
    parser.add_argument("liste", type=Path, help="Path to Liste.xlsx")
    args = parser.parse_args()

    day_str = f"{args.day_dir.name}.2025"
    day = dt.datetime.strptime(day_str, "%d.%m.%Y").date()
    month_sheet = day.strftime("%B_%y").capitalize()
    month_sheet = f"{MONTH_MAP[day.month]}_{day.strftime('%y')}"

    morning = next(args.day_dir.glob("*7*.xlsx"))
    evening_file = next(args.day_dir.glob("*19*.xlsx"), None)
    target_date, morning_summary = load_calls(morning)
    evening_summary: Dict[str, Dict[str, int]] = {}
    if evening_file:
        _, evening_summary = load_calls(evening_file)
    update_liste(args.liste, month_sheet, target_date, morning_summary, evening_summary)


if __name__ == "__main__":
    main()
