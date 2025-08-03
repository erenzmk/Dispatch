"""Utilities for resolving technician name aliases.

This module contains a small helper :func:`canonical_name` that is used
throughout the project to normalise technician names.  In the real world the
source data is far from consistent – sometimes we only receive the first name,
other times the full name in ``"Lastname, Firstname (Extra)"`` form.  The
previous implementation performed a simple case-insensitive fuzzy match which
meant that entries like ``"Ahmad, Daniyal (Keskin)"`` were not matched to the
existing ``"Daniyal"`` entry in ``Liste.xlsx``.  As a consequence new rows were
added for these verbose names and the daily numbers ended up in the wrong
place.

To make the matching more robust we now strip auxiliary information like
parentheses and leading surnames before performing the fuzzy comparison.  This
mirrors how dispatchers refer to colleagues in the spreadsheet and keeps the
output tidy.
"""

from difflib import get_close_matches
import re
from pathlib import Path
from typing import Iterable

try:  # pragma: no cover - optional dependency for reading Liste.xlsx
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - handled gracefully when missing
    load_workbook = None  # type: ignore[assignment]

# Map alternate spellings or abbreviations to their canonical form.
# Extend this mapping as needed. Keys are treated case-insensitively.
ALIASES = {
    # "jon doe": "John Doe",
}

# Cached lower-case mapping of aliases to canonical names.  Tests and callers
# may mutate :data:`ALIASES`, so provide a helper to rebuild this cache when
# needed.
_ALIAS_MAP = {k.lower(): v for k, v in ALIASES.items()}


def refresh_alias_map() -> None:
    """Rebuild the cached alias mapping from :data:`ALIASES`."""
    global _ALIAS_MAP
    _ALIAS_MAP = {k.lower(): v for k, v in ALIASES.items()}


def load_aliases(liste: Path | str) -> None:
    """Populate :data:`ALIASES` from the ``Zuordnungen`` sheet of *liste*.

    The optional "Zuordnungen" worksheet in ``Liste.xlsx`` stores mappings of
    unknown technician names to their canonical counterparts as created by the
    :mod:`assign_gui` helper.  Each row contains the unknown name in column A and
    the desired canonical name in column B.  When present these assignments are
    merged into :data:`ALIASES` so subsequent calls to :func:`canonical_name`
    resolve them automatically.

    Missing dependencies or absent worksheets are silently ignored to keep the
    calling code simple.
    """

    if load_workbook is None:
        return

    path = Path(liste)
    if not path.exists():
        return

    try:
        wb = load_workbook(path, read_only=True)
    except Exception:  # pragma: no cover - invalid workbook or other errors
        return

    try:
        if "Zuordnungen" not in wb.sheetnames:
            return
        ws = wb["Zuordnungen"]
        for unknown, known in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if unknown and known:
                ALIASES[str(unknown).strip()] = str(known).strip()
    finally:
        wb.close()

    refresh_alias_map()


def canonical_name(name: str, valid_names: Iterable[str], cutoff: float = 0.8) -> str:
    """Return the canonical representation for *name*.

    ``valid_names`` should contain the canonical names available in
    ``Liste.xlsx``.  First the static :data:`ALIASES` mapping is checked,
    otherwise :func:`difflib.get_close_matches` is used to find the closest
    match above ``cutoff``. Matching is performed case-insensitively so that
    names like ``"ALICE"`` still resolve to ``"Alice"``. Leading and trailing
    whitespace is removed before matching. If no match is found, the stripped
    name is returned unchanged.
    """

    # ``name`` may come in various formats, e.g. ``"Doe, John (Team)"`` or
    # just ``"john"``.  Normalise by removing parenthetical information and by
    # taking the part after a comma (which usually holds the first name).
    norm = name.strip()
    if not norm:
        return norm

    # Drop anything inside parentheses to ignore organisational hints.
    norm = re.sub(r"\([^)]*\)", "", norm)
    # If the remaining string contains a comma we assume ``Lastname, Firstname``
    # and keep the part after the last comma which typically is the name used in
    # the spreadsheet.
    if "," in norm:
        norm = norm.split(",")[-1]

    norm = norm.strip()
    key = norm.lower()

    if key in _ALIAS_MAP:
        return _ALIAS_MAP[key]

    valid_map = {v.lower(): v for v in valid_names}
    matches = get_close_matches(key, list(valid_map.keys()), n=1, cutoff=cutoff)
    return valid_map[matches[0]] if matches else norm
