"""Berichte nach Techniker-ID zusammenfassen."""

from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
from typing import List, Dict

import pandas as pd
from openpyxl import load_workbook

from .technicians import load_id_map


def summarize_report(excel_file: Path, liste_file: Path) -> List[Dict[str, object]]:
    """Bericht einlesen und neue/alte Calls je Techniker-ID zählen."""
    id_map = load_id_map(liste_file)
    valid_ids = set(id_map)

    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    file_date_raw = ws["A2"].value
    file_date = pd.to_datetime(file_date_raw, dayfirst=True).date()

    counts: Dict[str, Counter] = {}
    for row in ws.iter_rows(min_row=1, max_col=8, values_only=True):
        id_val = row[0]
        call_val = row[2] if len(row) > 2 else None
        open_val = row[7] if len(row) > 7 else None

        if id_val is None or open_val is None or call_val is None:
            continue
        id_str = str(id_val).strip()
        if id_str not in valid_ids:
            continue
        if not str(call_val).strip().startswith("17"):
            continue

        call_date = pd.to_datetime(open_val, dayfirst=True).date()
        diff = len(pd.bdate_range(call_date, file_date)) - 1
        status = "new" if diff <= 1 else "old"
        counts.setdefault(id_str, Counter())[status] += 1

    wb.close()

    results: List[Dict[str, object]] = []
    for id_str, counter in counts.items():
        new = counter.get("new", 0)
        old = counter.get("old", 0)
        results.append(
            {
                "id": id_str,
                "name": id_map.get(id_str, ""),
                "new": new,
                "old": old,
                "total": new + old,
            }
        )
    return results


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Zähle neue und alte Calls je Techniker-ID aus einem Bericht"
    )
    parser.add_argument("excel_file", type=Path, help="Pfad zur Excel-Datei")
    parser.add_argument("liste_file", type=Path, help="Pfad zur Technikerliste")
    parser.add_argument("--output", type=Path, help="Pfad für die Ausgabedatei (CSV)")
    args = parser.parse_args()

    summary = summarize_report(args.excel_file, args.liste_file)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(summary).to_csv(args.output, index=False)
        print(f"Ergebnis gespeichert in {args.output}")
    else:
        print(pd.DataFrame(summary).to_string(index=False))


if __name__ == "__main__":
    main()
