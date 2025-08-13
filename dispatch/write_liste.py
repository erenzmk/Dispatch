"""Trage Tagesdaten in ``Liste.xlsx`` ein.

Dieses Skript liest die täglichen Rohdaten aus ``data/reports`` ein und
schreibt die aggregierten Werte in das Monatsblatt ``Juli_25``.  Das Blatt
ist in Wochenblöcke organisiert; jeder Block besteht aus einem Spaltenpaar
``(Name, Datum)`` sowie weiteren Feldern rechts davon.  Für jeden Eintrag wird
die passende Zeile über den Techniker‑Namen und das Datum bestimmt.  Es werden
keine neuen Blöcke oder Zeilen angelegt und bestehende Inhalte werden nicht
überschrieben.
"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook

from .name_aliases import canonical_name, ALIASES, refresh_alias_map

INFILE = Path(r"C:\Users\egencer\Documents\GitHub\Dispatch\data\Liste.xlsx")

logger = logging.getLogger("write_liste")


def is_name_cell(v: object) -> bool:
    """Heuristik, ob ``v`` wie ein Name aussieht."""

    if v is None:
        return False
    s = str(v).strip()
    return bool(s) and not s.isdigit()


def is_date_cell(v: object) -> bool:
    """Prüfe, ob ``v`` ein Datum repräsentiert."""

    try:
        pd.to_datetime(v)
        return True
    except Exception:
        return False


def load_mapping(wb_path: Path) -> tuple[Dict[str, str], List[str]]:
    """Lese Alias-Mapping und Technikerliste."""

    df = pd.read_excel(wb_path, sheet_name="Technikernamen + PUDO", header=1)
    df = df[["first", "last", "dk"]].dropna(subset=["first"])
    df["name"] = (df["first"].fillna("") + " " + df["last"].fillna("")).str.strip()

    mapping: Dict[str, str] = {}
    for _, row in df.iterrows():
        alias = str(row.get("dk", "")).strip()
        name = row["name"]
        if alias and name:
            mapping[alias] = name

    tech_order = df["name"].tolist()
    return mapping, tech_order


def detect_week_blocks(ws, data_start_row: int = 2) -> List[dict]:
    """Suche alle Wochenblöcke im Arbeitsblatt.

    Ein Block besteht aus zwei Kernspalten (Name, Datum) sowie weiteren
    Feldern rechts davon.  Die Erkennung erfolgt anhand der Datenzeilen.
    """

    blocks: List[dict] = []
    max_col = ws.max_column
    max_row = ws.max_row
    col = 1
    while col < max_col:
        name_col = col
        date_col = col + 1
        if date_col > max_col:
            break
        hits = 0
        rows: List[int] = []
        for r in range(data_start_row, max_row + 1):
            nv = ws.cell(r, name_col).value
            dv = ws.cell(r, date_col).value
            if is_name_cell(nv) and is_date_cell(dv):
                hits += 1
                rows.append(r)
        if hits >= 5:
            # Woche bestimmen
            week = None
            for r in rows:
                dv = ws.cell(r, date_col).value
                if is_date_cell(dv):
                    d = pd.to_datetime(dv).date()
                    week = d.isocalendar()[1]
                    break
            # Feldzuordnung aus der Kopfzeile
            columns: Dict[str, int] = {}
            c = name_col
            while True:
                header = ws.cell(1, c).value
                if header is None:
                    break
                columns[str(header).strip().lower()] = c
                c += 1
            blocks.append(
                {
                    "name_col": name_col,
                    "date_col": date_col,
                    "rows": rows,
                    "week": week,
                    "columns": columns,
                }
            )
            col = c + 1  # nach dem Block (inkl. Trennspalte) fortfahren
        else:
            col += 1
    return blocks


def build_row_index(ws, block: dict, tech_order: List[str]) -> Dict[tuple, int]:
    """Erzeuge Index ``(name, datum) -> Zeile`` für ``block``."""

    idx: Dict[tuple, int] = {}
    name_col = block["name_col"]
    date_col = block["date_col"]
    for r in block["rows"]:
        nv = ws.cell(r, name_col).value
        dv = ws.cell(r, date_col).value
        if not (is_name_cell(nv) and is_date_cell(dv)):
            continue
        name = canonical_name(str(nv), tech_order)
        d = pd.to_datetime(dv).date()
        idx[(name.lower(), d.isoformat())] = r
    return idx


def aggregate_rows(rows: pd.DataFrame) -> dict[str, object]:
    """Fasse mehrere Zeilen nach den Regeln zusammen."""

    result: dict[str, object] = {}
    numeric = ["pre-closed", "total calls", "old calls", "new calls", "mails"]
    text = ["info", "details", "pudo"]

    for col in numeric:
        vals = pd.to_numeric(rows[col], errors="coerce")
        total = vals.sum(min_count=1)
        if pd.notna(total):
            result[col] = float(total)
    for col in text:
        parts = [str(v).strip() for v in rows[col].dropna() if str(v).strip()]
        if parts:
            result[col] = " | ".join(sorted(set(parts)))
    times = pd.to_datetime(rows["pickup time"], errors="coerce")
    if not times.isna().all():
        result["pickup time"] = times.min().time()
    valids = rows["valid"].dropna().astype(bool)
    if not valids.empty:
        result["valid"] = bool(valids.any())
    return result


def write_record(ws, block: dict, row_index: Dict[tuple, int], rec: dict) -> bool:
    """Schreibe ``rec`` in das Arbeitsblatt.

    ``rec`` muss mindestens ``name`` und ``date`` enthalten.  Rückgabe ist
    ``True`` bei Erfolg, sonst ``False``.
    """

    name = rec["name"]
    date = pd.to_datetime(rec["date"]).date()
    key = (name.lower(), date.isoformat())
    row = row_index.get(key)
    if not row:
        return False

    cols = block["columns"]
    ws.cell(row=row, column=cols.get("name"), value=name)
    ws.cell(row=row, column=cols.get("date"), value=date)
    if "weekday" in cols:
        ws.cell(row=row, column=cols["weekday"], value=date.strftime("%A"))

    for field, value in rec.items():
        if field.lower() in {"name", "date"}:
            continue
        col = cols.get(field.lower())
        if not col:
            continue
        if value in (None, ""):
            continue
        existing = ws.cell(row=row, column=col).value
        if existing in (None, ""):
            ws.cell(row=row, column=col, value=value)
    return True


def write_day(
    ws,
    blocks: List[dict],
    day_df: pd.DataFrame,
    date: dt.date,
    tech_order: List[str],
    normalize: bool = True,
) -> None:
    """Schreibe alle Daten des Tages ``date`` in ``ws``."""

    normalized = 0
    skipped = 0

    if "date" not in day_df.columns:
        day_df["date"] = date
    for idx, row in day_df.iterrows():
        if pd.isna(row.get("date")):
            day_df.at[idx, "date"] = date
            normalized += 1
            continue
        rdate = pd.to_datetime(row["date"]).date()
        if rdate != date:
            if normalize:
                day_df.at[idx, "date"] = date
                normalized += 1
            else:
                day_df.at[idx, "_skip"] = True
                skipped += 1
    if "_skip" in day_df.columns:
        day_df = day_df[~day_df["_skip"]]

    day_df["name"] = day_df["name"].map(lambda n: canonical_name(str(n), tech_order))
    grouped = day_df.groupby(["name", "date"], dropna=True)
    aggregated: Dict[str, dict[str, object]] = {}
    for (name, _), rows in grouped:
        aggregated[name] = aggregate_rows(rows)

    week = date.isocalendar()[1]
    block = next((b for b in blocks if b["week"] == week), None)
    if not block:
        logger.warning("Kein Wochenblock für %s gefunden", date.isoformat())
        return

    row_index = block.setdefault("row_index", build_row_index(ws, block, tech_order))

    for tech, values in aggregated.items():
        rec = {"name": tech, "date": date, **values}
        if not write_record(ws, block, row_index, rec):
            logger.warning("Keine Zeile für %s am %s", tech, date.isoformat())

    logger.info(
        "KW %s: %d Zeilen geschrieben, normalisiert %d, übersprungen %d",
        week,
        len(aggregated),
        normalized,
        skipped,
    )


def collect_day_df(day_dir: Path) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    for file in sorted(day_dir.iterdir()):
        if file.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            frames.append(pd.read_excel(file))
        elif file.suffix.lower() in {".csv", ".txt"}:
            frames.append(pd.read_csv(file, sep=";"))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--infile", default=str(INFILE))
    ap.add_argument("--out", default=None)
    ap.add_argument("--month", required=True)
    ap.add_argument("--inplace", action="store_true")
    ap.add_argument("--normalize-date", choices=["true", "false"], default="true")
    args = ap.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    wb = load_workbook(args.infile)
    ws = wb["Juli_25"]

    mapping, tech_order = load_mapping(Path(args.infile))
    ALIASES.update({k.lower(): v for k, v in mapping.items()})
    refresh_alias_map()

    blocks = detect_week_blocks(ws)
    for block in blocks:
        block["row_index"] = build_row_index(ws, block, tech_order)

    year, month = map(int, args.month.split("-"))
    last_day = (dt.date(year, month, 28) + dt.timedelta(days=4)).replace(day=1) - dt.timedelta(days=1)
    for day in range(1, last_day.day + 1):
        date = dt.date(year, month, day)
        day_dir = Path("data") / "reports" / f"{year:04d}-{month:02d}" / f"{day:02d}"
        if not day_dir.exists():
            continue
        day_df = collect_day_df(day_dir)
        if day_df.empty:
            continue
        write_day(
            ws,
            blocks,
            day_df,
            date,
            tech_order,
            normalize=args.normalize_date == "true",
        )

    out_file = args.infile if args.inplace or not args.out else args.out
    wb.save(out_file)


if __name__ == "__main__":
    main()

