from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook


def load_id_map(liste_path: Path) -> dict[str, str]:
    """Lese ein Mapping von Techniker-IDs zu Namen aus einer Excel-Datei.

    Erwartet eine Tabelle mit den Spalten ``ID`` und ``Techniker`` im ersten
    Arbeitsblatt. Whitespace wird entfernt und leere Zeilen werden ignoriert.
    """
    wb = load_workbook(liste_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_row = None
    headers: dict[str, int] = {}
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = {
            str(cell): col_idx
            for col_idx, cell in enumerate(row, start=1)
            if isinstance(cell, str)
        }
        if "ID" in headers and "Techniker" in headers:
            header_row = idx
            break

    result: dict[str, str] = {}
    if header_row is None:
        wb.close()
        return result

    id_col = headers["ID"]
    tech_col = headers["Techniker"]

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) < max(id_col, tech_col):
            continue
        id_value = row[id_col - 1]
        tech_value = row[tech_col - 1]
        if id_value is None or tech_value is None:
            continue
        key = str(id_value).strip()
        value = str(tech_value).strip()
        if key and value:
            result[key] = value

    wb.close()
    return result
