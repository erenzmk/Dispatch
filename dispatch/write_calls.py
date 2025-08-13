from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .name_aliases import canonical_name


def write_calls(workbook: Path | str, records: pd.DataFrame, sheet_name: str = "Juli") -> None:
    """Trage Call-Werte anhand von Namen und Datum in das Monatsblatt ein.

    ``records`` muss die Spalten ``name``, ``date`` und ``value`` enthalten.
    """
    wb = load_workbook(workbook)
    ws = wb[sheet_name]

    # Namen des ersten Blocks auslesen und auf Indizes abbilden
    name_map: dict[str, int] = {}
    valid_names: list[str] = []
    row_idx = 2
    while True:
        cell = ws.cell(row=row_idx, column=1)
        if not cell.value:
            break
        canon = canonical_name(str(cell.value).strip(), valid_names)
        if canon in name_map:
            break
        name_map[canon] = row_idx - 2
        valid_names.append(canon)
        row_idx += 1
    tech_rows = len(valid_names)

    for _, rec in records.iterrows():
        name = canonical_name(str(rec["name"]).strip(), valid_names)
        date = pd.to_datetime(rec["date"]).date()
        value = rec["value"]

        if name not in name_map:
            continue

        week_index = (date.day - 1) // 7
        call_col_idx = 2 + week_index * (13 + 1)
        call_col = get_column_letter(call_col_idx)
        day_offset = date.weekday()
        row = 2 + name_map[name] + tech_rows * day_offset
        ws[f"{call_col}{row}"] = value

    wb.save(workbook)
