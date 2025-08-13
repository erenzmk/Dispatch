import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from dispatch.write_calls import write_calls
from dispatch.name_aliases import refresh_alias_map


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli"
    names = ["Osama", "Daniyal", "Alice"]
    for idx, name in enumerate(names, start=2):
        ws.cell(row=idx, column=1, value=name)
    # Namen für weitere Tage kopieren
    for day in range(1, 7):
        for idx, name in enumerate(names, start=2 + len(names) * day):
            ws.cell(row=idx, column=1, value=name)
    wb.save(path)


def test_write_calls(tmp_path: Path) -> None:
    refresh_alias_map()
    file = tmp_path / "Juli.xlsx"
    _create_workbook(file)
    records = pd.DataFrame(
        [
            {"name": "Oussama", "date": dt.date(2025, 7, 1), "value": 5},
            {"name": "danyal", "date": dt.date(2025, 7, 15), "value": 3},
            {"name": "Alice", "date": dt.date(2025, 7, 29), "value": 7},
        ]
    )
    write_calls(file, records)
    wb = load_workbook(file)
    ws = wb["Juli"]
    assert ws["B5"].value == 5
    assert ws["AD6"].value == 3
    assert ws["BF7"].value == 7
