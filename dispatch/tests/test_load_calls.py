import datetime as dt
from pathlib import Path
import os

import pytest
from openpyxl import Workbook

from dispatch.process_reports import load_calls, RELEVANT_SHEET_PATTERNS


def test_load_calls_selects_correct_sheet(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1["A1"] = "not the sheet"

    ws2 = wb.create_sheet("Report")
    ws2["A2"] = dt.datetime(2025, 7, 1)
    ws2["A5"] = "Employee ID"
    ws2["B5"] = "Employee Name"
    ws2["C5"] = "Open Date Time"
    ws2["A6"] = 1
    ws2["B6"] = "Alice"
    ws2["C6"] = dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Alice": {"total": 1, "new": 1, "old": 0}}
    assert unknown == []


def test_load_calls_handles_west_central_sheet(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "West Central"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"], ws["B5"], ws["C5"] = (
        "Employee ID",
        "Employee Name",
        "Open Date Time",
    )
    ws["A6"], ws["B6"], ws["C6"] = 1, "Alice", dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Alice": {"total": 1, "new": 1, "old": 0}}
    assert unknown == []


def test_load_calls_resolves_alias(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"], ws["B5"], ws["C5"] = (
        "Employee ID",
        "Employee Name",
        "Open Date Time",
    )
    ws["A6"], ws["B6"], ws["C6"] = 1, "Oussama", dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path, ["Osama"])

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Osama": {"total": 1, "new": 1, "old": 0}}
    assert unknown == []


def test_load_calls_filters_irrelevant_sheets(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Meta"
    ws1["A5"], ws1["B5"], ws1["C5"] = (
        "Employee ID",
        "Employee Name",
        "Open Date Time",
    )
    ws1["A6"], ws1["B6"], ws1["C6"] = 1, "Bob", dt.datetime(2025, 6, 30)

    ws2 = wb.create_sheet("Report")
    ws2["A2"] = dt.datetime(2025, 7, 1)
    ws2["A5"], ws2["B5"], ws2["C5"] = (
        "Employee ID",
        "Employee Name",
        "Open Date Time",
    )
    ws2["A6"], ws2["B6"], ws2["C6"] = 1, "Alice", dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Alice": {"total": 1, "new": 1, "old": 0}}
    assert unknown == []


def test_load_calls_skips_duplicate_work_orders(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Report1"
    ws1["A2"] = dt.datetime(2025, 7, 1)
    ws1["A5"], ws1["B5"], ws1["C5"], ws1["D5"] = (
        "Employee ID",
        "Employee Name",
        "Work Order Number",
        "Open Date Time",
    )
    ws1["A6"], ws1["B6"], ws1["C6"], ws1["D6"] = (
        1,
        "Alice",
        "17500001",
        dt.datetime(2025, 6, 30),
    )

    ws2 = wb.create_sheet("Report2")
    ws2["A2"] = dt.datetime(2025, 7, 1)
    ws2["A5"], ws2["B5"], ws2["C5"], ws2["D5"] = (
        "Employee ID",
        "Employee Name",
        "Work Order Number",
        "Open Date Time",
    )
    ws2["A6"], ws2["B6"], ws2["C6"], ws2["D6"] = (
        2,
        "Bob",
        "17500001",
        dt.datetime(2025, 6, 29),
    )
    ws2["A7"], ws2["B7"], ws2["C7"], ws2["D7"] = (
        3,
        "Bob",
        "17500002",
        dt.datetime(2025, 6, 29),
    )

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {
        "Alice": {"total": 1, "new": 1, "old": 0},
        "Bob": {"total": 1, "new": 0, "old": 1},
    }
    assert unknown == []


def test_load_calls_without_employee_id(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = "Employee Name"
    ws["B5"] = "Open Date Time"
    ws["A6"] = "Alice"
    ws["B6"] = dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Alice": {"total": 1, "new": 1, "old": 0}}
    assert unknown == []


def test_load_calls_handles_header_variations(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = " employee id "
    ws["B5"] = "EMPLOYEE NAME "
    ws["C5"] = " open date time "
    ws["A6"] = 1
    ws["B6"] = "Alice"
    ws["C6"] = dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Alice": {"total": 1, "new": 1, "old": 0}}
    assert unknown == []


def test_load_calls_aggregates_all_sheets(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Report1"
    ws1["A2"] = dt.datetime(2025, 7, 1)
    ws1["A5"], ws1["B5"], ws1["C5"] = "Employee ID", "Employee Name", "Open Date Time"
    ws1["A6"], ws1["B6"], ws1["C6"] = 1, "Alice", dt.datetime(2025, 6, 30)

    ws2 = wb.create_sheet("Report2")
    ws2["A2"] = dt.datetime(2025, 7, 1)
    ws2["A5"], ws2["B5"], ws2["C5"] = "Employee ID", "Employee Name", "Open Date Time"
    ws2["A6"], ws2["B6"], ws2["C6"] = 2, "Bob", dt.datetime(2025, 6, 29)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {
        "Alice": {"total": 1, "new": 1, "old": 0},
        "Bob": {"total": 1, "new": 0, "old": 1},
    }
    assert unknown == []


def test_load_calls_missing_required_columns(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = "Employee ID"
    ws["B5"] = "Employee Name"  # "Open Date Time" fehlt komplett
    ws["A6"] = 1
    ws["B6"] = "Alice"

    path = tmp_path / "report.xlsx"
    wb.save(path)

    with pytest.raises(ValueError) as exc:
        load_calls(path)
    assert "Header row not found" in str(exc.value)

@pytest.mark.skipif(
    not Path("/proc/self/fd").exists(),
    reason="Requires /proc filesystem for file descriptor counting",
)
def test_load_calls_does_not_leak_file_handles(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"] = "Employee ID"
    ws["B5"] = "Employee Name"
    ws["C5"] = "Open Date Time"
    ws["A6"] = 1
    ws["B6"] = "Alice"
    ws["C6"] = dt.datetime(2025, 6, 30)

    path = tmp_path / "report.xlsx"
    wb.save(path)

    def fd_count() -> int:
        return len(os.listdir("/proc/self/fd"))

    baseline = fd_count()
    for _ in range(3):
        load_calls(path)
        assert fd_count() == baseline


def test_load_calls_ignores_non_call_numbers(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A2"] = dt.datetime(2025, 7, 1)
    ws["A5"], ws["B5"], ws["C5"], ws["D5"] = (
        "Employee ID",
        "Employee Name",
        "Work Order Number",
        "Open Date Time",
    )
    ws["A6"], ws["B6"], ws["C6"], ws["D6"] = 1, "Alice", "17500001", dt.datetime(2025, 6, 30)
    ws["A7"], ws["B7"], ws["C7"], ws["D7"] = 2, "Alice", "HOURS", dt.datetime(2025, 6, 30)
    path = tmp_path / "report.xlsx"
    wb.save(path)

    target_date, summary, unknown = load_calls(path)

    assert target_date == dt.date(2025, 7, 1)
    assert summary == {"Alice": {"total": 1, "new": 1, "old": 0}}
    assert unknown == []


def test_load_calls_reports_missing_relevant_sheets(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    path = tmp_path / "report.xlsx"
    wb.save(path)

    with pytest.raises(ValueError) as exc:
        load_calls(path)

    msg = str(exc.value)
    for pattern in RELEVANT_SHEET_PATTERNS:
        assert pattern.pattern in msg
    assert ws.title in msg
