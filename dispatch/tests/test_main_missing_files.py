import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
import datetime as dt

from dispatch.process_reports import main


def create_liste(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    wb.save(path)


def test_main_no_excel_files(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    liste = tmp_path / "Liste.xlsx"
    create_liste(liste)

    day_dir = tmp_path / "2025-07" / "01"
    day_dir.mkdir(parents=True)

    monkeypatch.setattr(sys, "argv", ["process_reports.py", str(day_dir), str(liste)])
    with pytest.raises(FileNotFoundError):
        main()


def test_main_missing_morning_file_uses_fallback(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    liste = tmp_path / "Liste.xlsx"
    create_liste(liste)

    day_dir = tmp_path / "2025-07" / "01"
    day_dir.mkdir(parents=True)

    # create an evening file to ensure only morning is missing
    wb = Workbook()
    fallback = day_dir / "evening19.xlsx"
    wb.save(fallback)

    monkeypatch.setattr(sys, "argv", ["process_reports.py", str(day_dir), str(liste)])

    called = {}

    def fake_load_calls(path, valid_names=None):
        called["used"] = Path(path)
        return dt.date(2025, 7, 1), {}, []

    def fake_update_liste(liste_path, month_sheet, target_date, morning_summary):
        pass

    monkeypatch.setattr("dispatch.process_reports.load_calls", fake_load_calls)
    monkeypatch.setattr("dispatch.process_reports.update_liste", fake_update_liste)

    main()

    assert called["used"] == fallback


def test_main_no_evening_file_ok(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    liste = tmp_path / "Liste.xlsx"
    create_liste(liste)

    day_dir = tmp_path / "2025-07" / "01"
    day_dir.mkdir(parents=True)

    # create a morning file but no evening file
    wb = Workbook()
    wb.save(day_dir / "morning7.xlsx")

    monkeypatch.setattr(sys, "argv", ["process_reports.py", str(day_dir), str(liste)])

    def fake_load_calls(path, valid_names=None):
        return dt.date(2025, 7, 1), {}, []

    def fake_update_liste(liste_path, month_sheet, target_date, morning_summary):
        pass

    monkeypatch.setattr("dispatch.process_reports.load_calls", fake_load_calls)
    monkeypatch.setattr("dispatch.process_reports.update_liste", fake_update_liste)

    # Should not raise even if evening file is missing
    main()


def test_main_custom_morning_pattern(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    liste = tmp_path / "Liste.xlsx"
    create_liste(liste)

    day_dir = tmp_path / "2025-07" / "01"
    day_dir.mkdir(parents=True)

    wb = Workbook()
    wb.save(day_dir / "report_morning.xlsx")

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "process_reports.py",
            str(day_dir),
            str(liste),
            "--morning-pattern",
            "*morning.xlsx",
        ],
    )

    def fake_load_calls(path, valid_names=None):
        return dt.date(2025, 7, 1), {}, []

    def fake_update_liste(liste_path, month_sheet, target_date, morning_summary):
        pass

    monkeypatch.setattr("dispatch.process_reports.load_calls", fake_load_calls)
    monkeypatch.setattr("dispatch.process_reports.update_liste", fake_update_liste)

    main()


def test_main_creates_missing_sheet(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    liste = tmp_path / "Liste.xlsx"
    wb = Workbook()
    wb.save(liste)

    day_dir = tmp_path / "2025-07" / "01"
    day_dir.mkdir(parents=True)
    wb = Workbook()
    wb.save(day_dir / "morning7.xlsx")

    monkeypatch.setattr(sys, "argv", ["process_reports.py", str(day_dir), str(liste)])

    def fake_load_calls(path, valid_names=None):
        return dt.date(2025, 7, 1), {"Alice": {"total": 1, "new": 1, "old": 0}}, []

    monkeypatch.setattr("dispatch.process_reports.load_calls", fake_load_calls)

    monkeypatch.setattr("builtins.input", lambda _: "j")
    main()

    wb2 = load_workbook(liste)
    assert "Juli_25" in wb2.sheetnames
    assert wb2["Juli_25"].max_row == 1
    wb2.close()


def test_main_selects_existing_sheet(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    liste = tmp_path / "Liste.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Juni_25"
    wb.save(liste)

    day_dir = tmp_path / "2025-07" / "01"
    day_dir.mkdir(parents=True)
    wb2 = Workbook()
    wb2.save(day_dir / "morning7.xlsx")

    monkeypatch.setattr(sys, "argv", ["process_reports.py", str(day_dir), str(liste)])

    def fake_load_calls(path, valid_names=None):
        return dt.date(2025, 7, 1), {}, []

    called = {}

    def fake_update_liste(liste_path, month_sheet, target_date, morning_summary):
        called["month_sheet"] = month_sheet

    monkeypatch.setattr("dispatch.process_reports.load_calls", fake_load_calls)
    monkeypatch.setattr("dispatch.process_reports.update_liste", fake_update_liste)
    monkeypatch.setattr("builtins.input", lambda _: "Juni_25")
    main()

    assert called["month_sheet"] == "Juni_25"
