import datetime as dt
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest
import logging

from dispatch.process_reports import update_liste, excel_to_date


def add_block_headers(ws, date_col: int) -> None:
    ws.cell(row=1, column=date_col, value="date")
    ws.cell(row=1, column=date_col + 1, value="weekday")
    ws.cell(row=1, column=date_col + 7, value="total calls")
    ws.cell(row=1, column=date_col + 8, value="old calls")
    ws.cell(row=1, column=date_col + 9, value="new calls")


def test_mismatched_date_skipped(tmp_path: Path, caplog: pytest.LogCaptureFixture) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 2))
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 0, "old": 1}}

    import dispatch.process_reports as pr

    logger = pr.logger
    original_handlers = logger.handlers[:]
    original_propagate = logger.propagate
    logger.handlers = []
    logger.propagate = True
    try:
        with caplog.at_level(logging.WARNING, logger="dispatch.process_reports"):
            update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)
    finally:
        logger.handlers = original_handlers
        logger.propagate = original_propagate

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 2)
    assert ws2.cell(row=2, column=10).value is None
    assert "Abweichende Datumsangabe" in caplog.text
    wb2.close()


def test_mismatched_date_fixed(tmp_path: Path, caplog: pytest.LogCaptureFixture) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 2))
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 0, "old": 1}}

    import dispatch.process_reports as pr

    logger = pr.logger
    original_handlers = logger.handlers[:]
    original_propagate = logger.propagate
    logger.handlers = []
    logger.propagate = True
    try:
        with caplog.at_level(logging.WARNING, logger="dispatch.process_reports"):
            update_liste(
                file,
                "Juli_25",
                dt.date(2025, 7, 1),
                morning,
                fix_mismatched_dates=True,
            )
    finally:
        logger.handlers = original_handlers
        logger.propagate = original_propagate

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=10).value == 1
    assert "Abweichende Datumsangabe" in caplog.text
    wb2.close()
