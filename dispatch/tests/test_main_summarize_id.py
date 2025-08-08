import ast
import subprocess
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


def test_cli_summarize_id(tmp_path: Path) -> None:
    liste = tmp_path / "Liste.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Techniker"])
    ws.append(["1", "Alice"])
    ws.append(["2", "Bob"])
    wb.save(liste)
    wb.close()

    report = tmp_path / "report.xlsx"
    data = pd.DataFrame(
        [
            ["1", "", "17500001", "", "", "", "", "30-06-2025 09:04"],
            ["1", "", "17500002", "", "", "", "", "26-06-2025 15:03"],
            ["2", "", "17500003", "", "", "", "", "30-06-2025 09:04"],
            ["3", "", "17500004", "", "", "", "", "30-06-2025 09:04"],
        ],
        columns=list("ABCDEFGH"),
    )
    with pd.ExcelWriter(report) as writer:
        data.to_excel(writer, index=False, header=False, startrow=2)
    wb = load_workbook(report)
    ws = wb.worksheets[0]
    ws["A2"] = "01.07.25 07:01"
    wb.save(report)
    wb.close()

    out_csv = tmp_path / "summary.csv"
    subprocess.run(
        [
            "python",
            "-m",
            "dispatch.main",
            "summarize-id",
            str(report),
            str(liste),
            "-o",
            str(out_csv),
        ],
        check=True,
    )

    df = pd.read_csv(out_csv)
    df["id"] = df["id"].astype(str)
    df["calls"] = df["calls"].apply(ast.literal_eval)
    df = df.sort_values("id").reset_index(drop=True)
    expected = pd.DataFrame(
        [
            {
                "id": "1",
                "name": "Alice",
                "new": 1,
                "old": 1,
                "total": 2,
                "calls": ["17500001", "17500002"],
            },
            {
                "id": "2",
                "name": "Bob",
                "new": 1,
                "old": 0,
                "total": 1,
                "calls": ["17500003"],
            },
        ]
    )
    pd.testing.assert_frame_equal(df, expected)

    out_csv.unlink()
    report.unlink()
    liste.unlink()
