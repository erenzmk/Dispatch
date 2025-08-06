import pandas as pd
from openpyxl import load_workbook, Workbook

from dispatch.summarize_by_id import summarize_report


def test_summarize_report_by_id(tmp_path):
    liste = tmp_path / "Liste.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Techniker"])
    ws.append(["1", "Alice"])
    ws.append(["2", "Bob"])
    wb.save(liste)
    wb.close()

    report = tmp_path / "report.xlsx"
    data = pd.DataFrame(
        [
            ["1", "", "17500001", "", "", "", "", "30-06-2025 09:04"],
            ["1", "", "17500002", "", "", "", "", "26-06-2025 15:03"],
            ["2", "", "17500003", "", "", "", "", "30-06-2025 09:04"],
            ["3", "", "17500004", "", "", "", "", "30-06-2025 09:04"],
        ],
        columns=list("ABCDEFGH"),
    )
    with pd.ExcelWriter(report) as writer:
        data.to_excel(writer, index=False, header=False, startrow=2)
    wb = load_workbook(report)
    ws = wb.worksheets[0]
    ws["A2"] = "01.07.25 07:01"
    wb.save(report)
    wb.close()

    result = summarize_report(report, liste)
    assert sorted(result, key=lambda r: r["id"]) == [
        {"id": "1", "name": "Alice", "new": 1, "old": 1, "total": 2},
        {"id": "2", "name": "Bob", "new": 1, "old": 0, "total": 1},
    ]
