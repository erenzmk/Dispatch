import datetime as dt
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest
import logging

from dispatch.process_reports import update_liste, excel_to_date, PREV_DAY_MAP
from dispatch.name_aliases import refresh_alias_map


def add_block_headers(ws, date_col: int, with_name: bool = False) -> None:
    if with_name:
        ws.cell(row=1, column=date_col - 1, value="name")
    ws.cell(row=1, column=date_col, value="date")
    ws.cell(row=1, column=date_col + 1, value="weekday")
    ws.cell(row=1, column=date_col + 7, value="total calls")
    ws.cell(row=1, column=date_col + 8, value="old calls")
    ws.cell(row=1, column=date_col + 9, value="new calls")


def test_update_liste(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 1))
    ws.cell(row=3, column=1, value="Bob")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 3, "new": 1, "old": 2}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]

    assert ws2.cell(row=2, column=10).value == 3
    assert ws2.cell(row=2, column=11).value == 2
    assert ws2.cell(row=2, column=12).value == 1
    assert ws2.cell(row=3, column=10).value is None


def test_update_liste_empty_morning(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    with pytest.raises(ValueError, match="no data"):
        update_liste(file, "Juli_25", dt.date(2025, 7, 1), {})


def test_update_liste_resolves_name_alias(tmp_path: Path, caplog):
    refresh_alias_map()
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Osama")
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 1))
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Oussama": {"total": 2, "new": 1, "old": 1}}

    import dispatch.process_reports as pr

    logger = pr.logger
    original_handlers = logger.handlers[:]
    original_propagate = logger.propagate
    logger.handlers = []
    logger.propagate = True
    try:
        with caplog.at_level(logging.WARNING, logger="dispatch.process_reports"):
            pr.update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)
    finally:
        logger.handlers = original_handlers
        logger.propagate = original_propagate

    assert "Techniker" not in caplog.text

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=2, column=10).value == 2
    assert ws2.cell(row=2, column=11).value == 1
    assert ws2.cell(row=2, column=12).value == 1
    wb2.close()


def test_update_liste_adds_missing_technician(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Oussama": {"total": 2, "new": 1, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=2, column=1).value == "Osama"
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=10).value == 2
    assert ws2.cell(row=2, column=11).value == 1
    assert ws2.cell(row=2, column=12).value == 1
    wb2.close()


def test_update_liste_sorts_rows(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    # unsortierte Ausgangsdaten
    ws.cell(row=2, column=1, value="Bob")
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 1))
    ws.cell(row=3, column=1, value="Alice")
    ws.cell(row=3, column=3, value=dt.date(2025, 7, 1))
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {
        "Bob": {"total": 2, "new": 1, "old": 1},
        "Alice": {"total": 1, "new": 0, "old": 1},
    }

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=2, column=1).value == "Alice"
    assert ws2.cell(row=2, column=10).value == 1
    assert ws2.cell(row=3, column=1).value == "Bob"
    assert ws2.cell(row=3, column=10).value == 2
    wb2.close()


def test_update_liste_uses_technician_header_column(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=3, value="Techniker")
    add_block_headers(ws, 5)
    ws.cell(row=2, column=3, value="Alice")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 0, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=2, column=3).value == "Alice"
    assert excel_to_date(ws2.cell(row=2, column=5).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=12).value == 1
    wb2.close()


def test_update_liste_multiple_runs(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    add_block_headers(ws, 17)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 1))
    ws.cell(row=2, column=17, value=dt.date(2025, 7, 2))
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 0, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)
    update_liste(file, "Juli_25", dt.date(2025, 7, 2), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=2, column=10).value == 1
    assert ws2.cell(row=2, column=24).value == 1
    wb2.close()


def test_update_liste_creates_missing_day_block(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Alice")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 0, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 2), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=1, column=16).value == "Name"
    assert excel_to_date(ws2.cell(row=2, column=17).value) == dt.date(2025, 7, 2)
    assert ws2.cell(row=2, column=24).value == 1
    assert ws2.cell(row=2, column=25).value == 1
    assert ws2.cell(row=2, column=26).value == 0
    wb2.close()


def test_update_liste_creates_first_day_block(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    ws.cell(row=2, column=1, value="Alice")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 1, "old": 0}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=1, column=2).value == "Name"
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=10).value == 1
    assert ws2.cell(row=2, column=11).value == 0
    assert ws2.cell(row=2, column=12).value == 1
    wb2.close()


def test_update_liste_day_blocks_with_blank_columns(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    add_block_headers(ws, 17)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=3, column=1, value="Bob")

    # Tag 1
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 1))
    ws.cell(row=2, column=10, value=3)
    ws.cell(row=3, column=3, value=dt.date(2025, 7, 1))
    ws.cell(row=3, column=10, value=4)

    # Tag 2
    ws.cell(row=2, column=17, value=dt.date(2025, 7, 2))
    ws.cell(row=2, column=24, value=5)
    ws.cell(row=3, column=17, value=dt.date(2025, 7, 2))
    ws.cell(row=3, column=24, value=6)

    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {
        "Alice": {"total": 7, "new": 3, "old": 4},
        "Bob": {"total": 2, "new": 1, "old": 1},
    }

    day = dt.date(2025, 7, 2)
    update_liste(file, "Juli_25", day, morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]

    assert ws2.cell(row=2, column=24).value == 7
    assert ws2.cell(row=3, column=24).value == 2
    assert excel_to_date(ws2.cell(row=2, column=17).value) == day
    assert excel_to_date(ws2.cell(row=3, column=17).value) == day
    assert ws2.max_row == 3
    assert ws2.max_column >= 24
    wb2.close()


def test_update_liste_keeps_existing_date(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=3, column=1, value="Alice")
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 2))
    ws.cell(row=3, column=3, value=dt.date(2025, 7, 1))
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 5, "new": 2, "old": 3}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    # Zeile mit abweichendem Datum bleibt unverändert
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 2)
    assert ws2.cell(row=2, column=10).value is None
    # Zeile mit passendem Datum wird beschrieben
    assert excel_to_date(ws2.cell(row=3, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=3, column=10).value == 5
    assert ws2.cell(row=3, column=11).value == 3
    assert ws2.cell(row=3, column=12).value == 2
    wb2.close()


def test_update_liste_appends_new_row_on_date_mismatch(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=3, value=dt.date(2025, 7, 2))
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 0, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    # Originale Zeile bleibt unverändert
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 2)
    assert ws2.cell(row=2, column=10).value is None
    # Neue Zeile mit aktuellem Datum wird angefügt
    assert excel_to_date(ws2.cell(row=3, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=3, column=10).value == 1
    assert ws2.cell(row=3, column=11).value == 1
    assert ws2.cell(row=3, column=12).value == 0
    wb2.close()


def test_update_liste_renames_name_column(tmp_path: Path):
    file = tmp_path / "liste.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Name")
    add_block_headers(ws, 3)
    wb.save(file)

    morning = {"Alice": {"total": 2, "new": 1, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=1, column=1).value == "Techniker"
    assert ws2.cell(row=2, column=1).value == "Alice"
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=10).value == 2
    assert ws2.cell(row=2, column=11).value == 1
    assert ws2.cell(row=2, column=12).value == 1
    wb2.close()


def test_update_liste_creates_missing_sheet(tmp_path: Path):
    file = tmp_path / "liste.xlsx"
    wb = Workbook()
    wb.save(file)

    morning = {"Alice": {"total": 2, "new": 1, "old": 1}}

    with pytest.raises(ValueError):
        update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)


def test_update_liste_handles_name_column_in_day_block(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 4, with_name=True)
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 2, "new": 1, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert excel_to_date(ws2.cell(row=2, column=4).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=11).value == 2
    assert ws2.cell(row=2, column=12).value == 1
    assert ws2.cell(row=2, column=13).value == 1
    wb2.close()


def test_update_liste_accepts_weekday_in_date_column(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    ws.cell(row=1, column=3, value="weekday")
    ws.cell(row=1, column=4, value="Wochentag")
    ws.cell(row=1, column=10, value="total calls")
    ws.cell(row=1, column=11, value="old calls")
    ws.cell(row=1, column=12, value="new calls")
    ws.cell(row=2, column=1, value="Alice")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 2, "new": 1, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    # Datum wird trotz falsch benannter Kopfzeile korrekt eingetragen
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=4).value == "Dienstag"
    assert ws2.cell(row=2, column=10).value == 2
    assert ws2.cell(row=2, column=11).value == 1
    assert ws2.cell(row=2, column=12).value == 1
    wb2.close()


def test_update_liste_handles_invalid_header(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    ws.cell(row=1, column=3, value="Falsch")
    ws.cell(row=1, column=10, value="total calls")
    ws.cell(row=1, column=11, value="old calls")
    ws.cell(row=1, column=12, value="new calls")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 0, "old": 1}}

    update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=10).value == 1
    wb2.close()


def test_excel_to_date_none():
    with pytest.raises(ValueError, match="Leere Zelle"):
        excel_to_date(None)


def test_excel_to_date_invalid():
    with pytest.raises(ValueError, match="Ungültiger Datumswert"):
        excel_to_date("abc")


@pytest.mark.parametrize(
    "date_value",
    [None, "abc", '=TEXT(B3,"tttt")', "Mittwoch"],
)
def test_update_liste_fills_invalid_date_cell(tmp_path: Path, caplog, date_value):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Alice")
    if date_value is not None:
        ws.cell(row=2, column=3, value=date_value)
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 1, "new": 0, "old": 1}}

    import dispatch.process_reports as pr

    logger = pr.logger
    original_handlers = logger.handlers[:]
    original_propagate = logger.propagate
    logger.handlers = []
    logger.propagate = True
    try:
        with caplog.at_level(logging.WARNING, logger="dispatch.process_reports"):
            update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)
    finally:
        logger.handlers = original_handlers
        logger.propagate = original_propagate

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert excel_to_date(ws2.cell(row=2, column=3).value) == dt.date(2025, 7, 1)
    assert ws2.cell(row=2, column=10).value == 1
    assert ws2.cell(row=2, column=11).value == 1
    assert ws2.cell(row=2, column=12).value == 0
    wb2.close()

    assert sum(1 for r in caplog.records if r.levelno == logging.WARNING) == 1


def test_update_liste_prev_day_after_total(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    ws.cell(row=1, column=3, value="date")
    ws.cell(row=1, column=10, value="total calls")
    ws.cell(row=1, column=11, value="old calls")
    ws.cell(row=1, column=12, value="new calls")
    ws.cell(row=1, column=13, value="weekday")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {"Alice": {"total": 2, "new": 1, "old": 1}}
    day = dt.date(2025, 7, 1)

    update_liste(file, "Juli_25", day, morning)

    wb2 = load_workbook(file)
    ws2 = wb2["Juli_25"]
    assert ws2.cell(row=2, column=10).value == 2
    assert ws2.cell(row=2, column=11).value == 1
    assert ws2.cell(row=2, column=12).value == 1
    assert ws2.cell(row=2, column=13).value == PREV_DAY_MAP[day.weekday()]
    assert ws2.max_column >= 13
    wb2.close()


def test_warning_summary_aggregates(tmp_path: Path, caplog):
    wb = Workbook()
    ws = wb.active
    ws.title = "Juli_25"
    ws.cell(row=1, column=1, value="Techniker")
    add_block_headers(ws, 3)
    ws.cell(row=2, column=1, value="Alice")
    ws.cell(row=2, column=3, value="abc")
    ws.cell(row=3, column=1, value="Bob")
    ws.cell(row=3, column=3, value="xyz")
    file = tmp_path / "liste.xlsx"
    wb.save(file)

    morning = {
        "Alice": {"total": 1, "new": 0, "old": 1},
        "Bob": {"total": 2, "new": 1, "old": 1},
    }

    import dispatch.process_reports as pr

    logger = pr.logger
    original_handlers = logger.handlers[:]
    original_propagate = logger.propagate
    logger.handlers = []
    logger.propagate = True
    try:
        with caplog.at_level(logging.INFO):
            update_liste(file, "Juli_25", dt.date(2025, 7, 1), morning)
    finally:
        logger.handlers = original_handlers
        logger.propagate = original_propagate

    assert "2 ungültige Datumsangaben automatisch korrigiert" in caplog.text
